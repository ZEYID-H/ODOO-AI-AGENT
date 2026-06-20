"""Customer statement export (CSV / Excel).

Additive utility built ON TOP OF get_customer_statement. It does not modify any
existing tool, schema, provider, or routing. It is NOT an OpenAI-routed tool —
it is invoked by the UI to produce a downloadable file.

CSV is mandatory; XLSX requires openpyxl (raised clearly if absent).
"""

import io
import csv
import re
from datetime import date

from src.tools.customer_tools import get_customer_statement

CURRENCY = "QAR"


def _safe_filename(customer_name: str, ext: str) -> str:
    slug = re.sub(r"[^A-Za-z0-9]+", "_", customer_name).strip("_")
    return f"{slug}_Statement_{date.today().isoformat()}.{ext}"


def _summary_rows(data: dict) -> list[tuple]:
    return [
        ("Total Invoiced", round(data["total_invoiced"], 2)),
        ("Total Paid", round(data["total_paid"], 2)),
        ("Outstanding Balance", round(data["outstanding_balance"], 2)),
        ("Invoice Count", data["invoice_count"]),
        ("Payment Count", data["payment_count"]),
    ]


def _reconciliation_note(data: dict) -> str:
    if data["reconciles"]:
        return "Activity balance matches open-item balance."
    return (
        f"Activity balance {round(data['activity_balance'], 2)} vs open-item balance "
        f"{round(data['outstanding_balance'], 2)}; difference {round(data['difference'], 2)} "
        f"(unreconciled/advance payments)."
    )


def _statement_to_csv(data: dict) -> str:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Customer Statement"])
    w.writerow(["Customer", data["customer_name"]])
    w.writerow(["Statement Date", date.today().isoformat()])
    w.writerow(["Currency", CURRENCY])
    w.writerow([])

    w.writerow(["Summary"])
    for label, val in _summary_rows(data):
        w.writerow([label, val])
    w.writerow([])

    w.writerow(["Transactions"])
    w.writerow(["Date", "Type", "Reference", "Debit", "Credit", "Running Balance"])
    for r in data["rows"]:
        w.writerow([r["date"], r["type"], r["reference"],
                    round(r["debit"], 2), round(r["credit"], 2), round(r["balance"], 2)])
    w.writerow([])

    w.writerow(["Reconciliation"])
    w.writerow(["Reconciled", "Yes" if data["reconciles"] else "No"])
    w.writerow(["Difference", round(data["difference"], 2)])
    w.writerow(["Note", _reconciliation_note(data)])
    return buf.getvalue()


def _statement_to_xlsx(data: dict) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as e:
        raise RuntimeError(
            "Excel export requires openpyxl. Install with: pip install openpyxl"
        ) from e

    wb = Workbook()
    ws = wb.active
    ws.title = "Statement"
    bold = Font(bold=True)

    def add(values, bold_header=False):
        ws.append(values)
        if bold_header:
            for col in range(1, len(values) + 1):
                ws.cell(row=ws.max_row, column=col).font = bold

    add(["Customer Statement"], bold_header=True)
    add(["Customer", data["customer_name"]])
    add(["Statement Date", date.today().isoformat()])
    add(["Currency", CURRENCY])
    add([])

    add(["Summary"], bold_header=True)
    for label, val in _summary_rows(data):
        add([label, val])
    add([])

    add(["Transactions"], bold_header=True)
    add(["Date", "Type", "Reference", "Debit", "Credit", "Running Balance"], bold_header=True)
    for r in data["rows"]:
        add([r["date"], r["type"], r["reference"],
             round(r["debit"], 2), round(r["credit"], 2), round(r["balance"], 2)])
    add([])

    add(["Reconciliation"], bold_header=True)
    add(["Reconciled", "Yes" if data["reconciles"] else "No"])
    add(["Difference", round(data["difference"], 2)])
    add(["Note", _reconciliation_note(data)])

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def export_customer_statement(customer_name: str, format: str = "csv") -> dict:
    """Generate a downloadable statement. Returns {filename, mimetype, content, format}
    or {error}. content is str for CSV, bytes for XLSX. Totals come from the full
    statement; no existing tool is modified."""
    data = get_customer_statement(customer_name)
    if "error" in data:
        return {"error": data["error"]}

    fmt = (format or "csv").lower()
    if fmt == "csv":
        return {
            "filename": _safe_filename(data["customer_name"], "csv"),
            "mimetype": "text/csv",
            "content": _statement_to_csv(data),
            "format": "csv",
        }
    if fmt in ("xlsx", "excel"):
        return {
            "filename": _safe_filename(data["customer_name"], "xlsx"),
            "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "content": _statement_to_xlsx(data),
            "format": "xlsx",
        }
    return {"error": f"Unsupported format '{format}'. Use 'csv' or 'xlsx'."}
